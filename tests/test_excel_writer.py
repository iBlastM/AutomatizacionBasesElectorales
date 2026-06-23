from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from src.catalogos import construir_catalogo_dfdl
from src.configuracion import cargar_configuracion_anual
from src.excel_writer import escribir_xlsx
from src.formateador import FormateadorElectoral
from src.lector_origen import leer_tabla_principal


def _resultado_2018(data_dir: Path, origen_2018: Path):
    config = cargar_configuracion_anual("2018", data_dir)
    catalogo = construir_catalogo_dfdl(config)
    tabla, _ = leer_tabla_principal(origen_2018, config)
    resultado = FormateadorElectoral(config, catalogo).formatear(tabla)
    return config, resultado.df_base.head(3)


def _columna_por_encabezado(ws, encabezado: str) -> int:
    for col in range(1, ws.max_column + 1):
        if ws.cell(1, col).value == encabezado:
            return col
    raise AssertionError(f"No existe encabezado {encabezado}")


def test_escribe_encabezados_visibles_y_formulas(data_dir: Path, origen_2018: Path):
    config, df_base = _resultado_2018(data_dir, origen_2018)
    datos = escribir_xlsx(df_base, config)
    wb = load_workbook(BytesIO(datos), data_only=False)
    ws = wb["Formato"]

    encabezados = [ws.cell(1, col).value for col in range(1, len(config.encabezados_visibles) + 1)]
    assert encabezados == config.encabezados_visibles
    assert ws["K2"].value.startswith("=")
    assert ws["L2"].value.startswith("=")
    assert ws["M2"].value.startswith("=")
    assert ws["N2"].value.startswith("=")
    assert ws.cell(2, _columna_por_encabezado(ws, "TOT_VOTOS")).value.startswith("=")
    assert ws.cell(2, _columna_por_encabezado(ws, "VALIDACION")).value.startswith("=")


def test_escribe_hoja_calculos_oculta(data_dir: Path, origen_2018: Path):
    config, df_base = _resultado_2018(data_dir, origen_2018)
    datos = escribir_xlsx(df_base, config)
    wb = load_workbook(BytesIO(datos), data_only=False)

    assert "_calculos" in wb.sheetnames
    assert wb["_calculos"].sheet_state == "hidden"
    assert wb["Formato"].freeze_panes == "A2"
    assert wb["Formato"].auto_filter.ref is not None
