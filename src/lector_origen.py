from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .configuracion import ConfiguracionAnual


LISTA_NOMINAL_ALIASES = {"LISTA_NOMINAL_CASILLA", "LISTA_NOMINAL"}
TOTAL_VOTOS_ALIASES = {"TOTAL_VOTOS", "VOTOS_EMITIDOS"}
NULOS_ALIASES = {"NUM_VOTOS_NULOS", "VOTOS_NULOS", "NULOS"}


@dataclass(frozen=True)
class MetadataOrigen:
    fila_encabezado: int
    hoja: str
    filas_leidas: int


def _normalizar_columna(valor: object) -> str:
    return str(valor).strip().upper() if valor is not None else ""


def detectar_fila_encabezado(ruta_excel: Path, config: ConfiguracionAnual) -> int:
    wb = load_workbook(ruta_excel, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    partidos = {p.upper() for p in config.partidos}

    for numero_fila, row in enumerate(ws.iter_rows(values_only=True), start=1):
        valores = {_normalizar_columna(v) for v in row if v is not None}
        coincidencias_partidos = len(valores & partidos)
        if (
            "SECCION" in valores
            and valores & LISTA_NOMINAL_ALIASES
            and valores & TOTAL_VOTOS_ALIASES
            and valores & NULOS_ALIASES
            and coincidencias_partidos >= 2
        ):
            return numero_fila

    raise ValueError("No se detectó la tabla principal en la primera hoja.")


def leer_tabla_principal(ruta_excel: Path, config: ConfiguracionAnual) -> tuple[pd.DataFrame, MetadataOrigen]:
    fila_encabezado = detectar_fila_encabezado(ruta_excel, config)
    wb = load_workbook(ruta_excel, read_only=True, data_only=True)
    hoja = wb.sheetnames[0]

    df = pd.read_excel(ruta_excel, sheet_name=0, header=fila_encabezado - 1)
    df = df.dropna(axis=1, how="all")
    df.columns = [str(col).strip() for col in df.columns]

    if "SECCION" not in df.columns:
        raise ValueError("La tabla detectada no contiene SECCION.")

    df["SECCION"] = pd.to_numeric(df["SECCION"], errors="coerce")
    df = df.dropna(subset=["SECCION"]).copy()
    df["SECCION"] = df["SECCION"].astype(int)
    df = df.dropna(how="all")

    return df, MetadataOrigen(fila_encabezado=fila_encabezado, hoja=hoja, filas_leidas=len(df))
