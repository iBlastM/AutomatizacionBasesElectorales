from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .perfiles import PerfilFormato, normalizar_clave


@dataclass(frozen=True)
class MetadataPerfil:
    fila_encabezado: int
    hoja: str
    filas_leidas: int
    encoding: str | None = None


def leer_tabla_perfil(ruta: Path, perfil: PerfilFormato) -> tuple[pd.DataFrame, MetadataPerfil]:
    if ruta.suffix.lower() == ".csv":
        return _leer_csv(ruta)
    if ruta.suffix.lower() in {".xlsx", ".xlsm"}:
        return _leer_excel(ruta, perfil)
    raise ValueError("El archivo origen debe ser .xlsx, .xlsm o .csv.")


def _leer_csv(ruta: Path) -> tuple[pd.DataFrame, MetadataPerfil]:
    ultimo_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp1252", "latin1"):
        try:
            df = pd.read_csv(ruta, encoding=encoding)
            df = _limpiar_dataframe(df)
            return df, MetadataPerfil(fila_encabezado=1, hoja="CSV", filas_leidas=len(df), encoding=encoding)
        except UnicodeDecodeError as exc:
            ultimo_error = exc
    raise ValueError(f"No se pudo leer el CSV con codificaciones comunes: {ultimo_error}")


def _leer_excel(ruta: Path, perfil: PerfilFormato) -> tuple[pd.DataFrame, MetadataPerfil]:
    fila, hoja = _detectar_tabla_excel(ruta, perfil)
    df = pd.read_excel(ruta, sheet_name=hoja, header=fila - 1)
    df = _limpiar_dataframe(df)
    return df, MetadataPerfil(fila_encabezado=fila, hoja=hoja, filas_leidas=len(df))


def _detectar_tabla_excel(ruta: Path, perfil: PerfilFormato) -> tuple[int, str]:
    wb = load_workbook(ruta, read_only=True, data_only=True)
    mejor: tuple[int, int, str] | None = None

    for ws in wb.worksheets:
        for numero_fila, row in enumerate(ws.iter_rows(values_only=True), start=1):
            valores = [normalizar_clave(valor) for valor in row if valor is not None and str(valor).strip()]
            if not valores:
                continue
            puntaje = _puntuar_encabezado(valores, ws.title, perfil)
            if puntaje <= 0:
                continue
            candidato = (puntaje, numero_fila, ws.title)
            if mejor is None or candidato[0] > mejor[0]:
                mejor = candidato

    if mejor is None:
        raise ValueError("No se detectó la tabla principal en el archivo.")
    _, fila, hoja = mejor
    return fila, hoja


def _puntuar_encabezado(valores: list[str], hoja: str, perfil: PerfilFormato) -> int:
    conjunto = set(valores)
    puntaje = 0

    requeridos = {
        "seccion": perfil.aliases_columnas.get("seccion", ["SECCION"]),
        "lista": perfil.aliases_columnas.get("lista", ["LISTA_NOMINAL", "LISTA_NOMINAL_CASILLA"]),
        "votos": perfil.aliases_columnas.get("votos", ["TOTAL_VOTOS", "VOTOS_EMITIDOS"]),
        "nulos": perfil.aliases_columnas.get("nulos", ["NULOS", "NUM_VOTOS_NULOS"]),
        "municipio": perfil.aliases_columnas.get("municipio", ["MUNICIPIO", "MUNICIPIO_LOCAL"]),
    }
    for aliases in requeridos.values():
        if conjunto & {normalizar_clave(alias) for alias in aliases}:
            puntaje += 10

    if "SECCION" not in conjunto:
        return 0

    for partido in perfil.partidos_salida:
        aliases = perfil.aliases_columnas.get(partido, [partido])
        if conjunto & {normalizar_clave(alias) for alias in aliases}:
            puntaje += 2

    for _, aliases in perfil.ranking_grupos:
        if conjunto & {normalizar_clave(alias) for alias in aliases}:
            puntaje += 1

    hoja_norm = normalizar_clave(hoja)
    if hoja_norm == "SECCION" or hoja_norm.endswith("_SECCION"):
        puntaje += 30
    if "CASILLA" in hoja_norm:
        puntaje -= 20
    return puntaje


def _limpiar_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(axis=1, how="all").dropna(how="all").copy()
    df.columns = [str(col).strip() for col in df.columns]

    seccion_col = _resolver_columna(df, ["SECCION", "SECCIÓN", "Seccion"])
    if seccion_col is None:
        raise ValueError("La tabla detectada no contiene SECCION.")
    if seccion_col != "SECCION":
        df = df.rename(columns={seccion_col: "SECCION"})

    df["SECCION"] = pd.to_numeric(df["SECCION"], errors="coerce")
    df = df.dropna(subset=["SECCION"]).copy()
    df["SECCION"] = df["SECCION"].astype(int)
    return df.reset_index(drop=True)


def _resolver_columna(df: pd.DataFrame, aliases: list[str]) -> str | None:
    indice = {normalizar_clave(col): col for col in df.columns}
    for alias in aliases:
        col = indice.get(normalizar_clave(alias))
        if col is not None:
            return col
    return None
