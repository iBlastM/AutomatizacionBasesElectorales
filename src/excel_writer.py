from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .configuracion import ConfiguracionAnual


def escribir_xlsx(df_base: pd.DataFrame, config: ConfiguracionAnual) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato"
    calc = wb.create_sheet("_calculos")
    calc.sheet_state = "hidden"

    _escribir_encabezados(ws, config)
    _escribir_calculos(calc, df_base, config)
    _escribir_datos_y_formulas(ws, df_base, config)
    _aplicar_estilos(ws, config, len(df_base))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def escribir_dataframe_xlsx(df_base: pd.DataFrame, sheet_name: str = "Formato") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col_idx, header in enumerate(df_base.columns, start=1):
        ws.cell(1, col_idx, header)
    for row_idx, (_, row) in enumerate(df_base.iterrows(), start=2):
        for col_idx, header in enumerate(df_base.columns, start=1):
            ws.cell(row_idx, col_idx, _numero(row.get(header, "")))

    _aplicar_estilos_dataframe(ws, len(df_base), len(df_base.columns))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _escribir_encabezados(ws, config: ConfiguracionAnual) -> None:
    for col_idx, header in enumerate(config.encabezados_visibles, start=1):
        ws.cell(1, col_idx, header)


def _escribir_calculos(calc, df_base: pd.DataFrame, config: ConfiguracionAnual) -> None:
    for idx, partido in enumerate(config.partidos, start=1):
        calc.cell(1, idx, partido)
    for row_idx, (_, row) in enumerate(df_base.iterrows(), start=2):
        for col_idx, partido in enumerate(config.partidos, start=1):
            calc.cell(row_idx, col_idx, _numero(row.get(partido, 0)))


def _escribir_datos_y_formulas(ws, df_base: pd.DataFrame, config: ConfiguracionAnual) -> None:
    posiciones = _posiciones(config)
    emitted_col = posiciones["VOTOS_EMITIDOS"]
    nominal_col = posiciones["LISTA_NOMINAL"]
    party_indices = list(range(config.indice_inicio_partidos + 1, config.indice_nulos + 1, 2))
    pcn_indices = [idx + 1 for idx in party_indices]
    nulos_col = config.indice_nulos + 1

    for excel_row, (_, row) in enumerate(df_base.iterrows(), start=2):
        for col_idx, header in enumerate(config.encabezados_visibles, start=1):
            cell = ws.cell(excel_row, col_idx)
            if header in df_base.columns and header not in {"PCN", "VOTOS"}:
                cell.value = _numero(row.get(header, ""))
                continue
            cell.value = _formula_para_columna(
                header=header,
                col_idx=col_idx,
                row_idx=excel_row,
                posiciones=posiciones,
                party_indices=party_indices,
                pcn_indices=pcn_indices,
                emitted_col=emitted_col,
                nominal_col=nominal_col,
                nulos_col=nulos_col,
                total_partidos=len(config.partidos),
            )


def _formula_para_columna(
    header: str,
    col_idx: int,
    row_idx: int,
    posiciones: dict[str, int],
    party_indices: list[int],
    pcn_indices: list[int],
    emitted_col: int,
    nominal_col: int,
    nulos_col: int,
    total_partidos: int,
) -> Any:
    emitted = f"{get_column_letter(emitted_col)}{row_idx}"
    nominal = f"{get_column_letter(nominal_col)}{row_idx}"
    last_calc_col = get_column_letter(total_partidos)
    calc_votes = f"_calculos!$A{row_idx}:${last_calc_col}{row_idx}"
    calc_heads = f"_calculos!$A$1:${last_calc_col}$1"

    top_map = {
        "1ER_LUGAR": 1,
        "2DO_LUGAR": 2,
        "3ER_LUGAR": 3,
        "1PP_MV": 1,
        "2PP_MV": 2,
        "3PP_MV": 3,
    }
    vote_map = {
        "1ERO_VOTOS": 1,
        "2DO_VOTOS": 2,
        "3RO_VOTOS": 3,
    }
    if header in top_map:
        n = top_map[header]
        return f'=IFERROR(INDEX({calc_heads},1,MATCH(LARGE({calc_votes},{n}),{calc_votes},0)),"")'
    if header in vote_map:
        return f"=IFERROR(LARGE({calc_votes},{vote_map[header]}),0)"
    if header == "PARTICIPACION":
        return f"=IFERROR({emitted}/{nominal},0)"
    if header == "ABSTENCION":
        return f"=IFERROR(1-{get_column_letter(posiciones['PARTICIPACION'])}{row_idx},0)"
    if header == "DIF_VOTOS_2DO":
        return (
            f"=IFERROR({get_column_letter(posiciones['1ERO_VOTOS'])}{row_idx}"
            f"-{get_column_letter(posiciones['2DO_VOTOS'])}{row_idx},0)"
        )
    if header == "DIF_VOTOS_3RO":
        return (
            f"=IFERROR({get_column_letter(posiciones['2DO_VOTOS'])}{row_idx}"
            f"-{get_column_letter(posiciones['3RO_VOTOS'])}{row_idx},0)"
        )
    if header == "DIF_PCN_2DO":
        return f"=IFERROR({get_column_letter(posiciones['DIF_VOTOS_2DO'])}{row_idx}/{emitted},0)"
    if header == "DIF_PCN_3RO":
        return f"=IFERROR({get_column_letter(posiciones['DIF_VOTOS_3RO'])}{row_idx}/{emitted},0)"
    if header == "DIF_2DO":
        return (
            f"=IFERROR({get_column_letter(posiciones['VOTOS'])}{row_idx}"
            f"-{get_column_letter(posiciones['VOTOS.1'])}{row_idx},0)"
        )
    if header == "DIF_3RO":
        return (
            f"=IFERROR({get_column_letter(posiciones['VOTOS.1'])}{row_idx}"
            f"-{get_column_letter(posiciones['VOTOS.2'])}{row_idx},0)"
        )
    if header == "TOT_VOTOS":
        votos = ",".join(f"{get_column_letter(i)}{row_idx}" for i in party_indices + [nulos_col])
        return f"=SUM({votos})"
    if header == "VALIDACION":
        pcns = ",".join(f"{get_column_letter(i)}{row_idx}" for i in pcn_indices + [nulos_col + 1])
        return f"=SUM({pcns})"
    if header == "PCN":
        vote_col = get_column_letter(col_idx - 1)
        return f"=IFERROR({vote_col}{row_idx}/{emitted},0)"
    if header == "VOTOS":
        ordinal = _ordinal_votos(col_idx, posiciones)
        return f"=IFERROR(LARGE({calc_votes},{ordinal}),0)"
    return ""


def _ordinal_votos(col_idx: int, posiciones: dict[str, int]) -> int:
    ordered = [posiciones["VOTOS"], posiciones["VOTOS.1"], posiciones["VOTOS.2"]]
    return ordered.index(col_idx) + 1 if col_idx in ordered else 1


def _posiciones(config: ConfiguracionAnual) -> dict[str, int]:
    posiciones: dict[str, int] = {}
    conteos: dict[str, int] = {}
    for idx, header in enumerate(config.encabezados_visibles, start=1):
        ocurrencia = conteos.get(header, 0)
        key = header if ocurrencia == 0 else f"{header}.{ocurrencia}"
        posiciones[key] = idx
        conteos[header] = ocurrencia + 1
    return posiciones


def _numero(value: Any) -> Any:
    if pd.isna(value):
        return ""
    if isinstance(value, str):
        return value
    try:
        if float(value).is_integer():
            return int(value)
    except (TypeError, ValueError):
        return value
    return value


def _aplicar_estilos(ws, config: ConfiguracionAnual, total_rows: int) -> None:
    ws.freeze_panes = "A2"
    max_col = len(config.encabezados_visibles)
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max(total_rows + 1, 1)}"
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "geo": PatternFill("solid", fgColor="D9EAF7"),
        "participacion": PatternFill("solid", fgColor="E2F0D9"),
        "top": PatternFill("solid", fgColor="FFF2CC"),
        "partidos": PatternFill("solid", fgColor="FCE4D6"),
        "validacion": PatternFill("solid", fgColor="E4DFEC"),
    }
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(1, col_idx)
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.fill = _fill_para_columna(col_idx, config, fills)
        width = min(max(len(str(cell.value)) + 2, 10), 22)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    decimal_headers = {"PCN", "PARTICIPACION", "ABSTENCION", "DIF_PCN_2DO", "DIF_PCN_3RO"}
    text_headers = {"ENTIDAD", "MUNICIPIO", "1ER_LUGAR", "2DO_LUGAR", "3ER_LUGAR", "1PP_MV", "2PP_MV", "3PP_MV"}
    fill_green = PatternFill("solid", fgColor="C6EFCE")
    fill_red = PatternFill("solid", fgColor="FFC7CE")
    for row in ws.iter_rows(min_row=2, max_row=total_rows + 1):
        for cell in row:
            header = ws.cell(1, cell.column).value
            cell.border = border
            if header == "VALIDACION":
                cell.number_format = "0.00%"
            elif header in decimal_headers:
                cell.number_format = "0.00"
            elif header not in text_headers:
                cell.number_format = "#,##0"

    validacion_col = None
    for col_idx in range(1, max_col + 1):
        if ws.cell(1, col_idx).value == "VALIDACION":
            validacion_col = col_idx
            break
    if validacion_col:
        for row_idx in range(2, total_rows + 2):
            cell = ws.cell(row_idx, validacion_col)
            cell.fill = fill_green
        from openpyxl.formatting.rule import CellIsRule
        col_letter = get_column_letter(validacion_col)
        rango = f"{col_letter}2:{col_letter}{total_rows + 1}"
        ws.conditional_formatting.add(
            rango,
            CellIsRule(operator="notEqual", formula=["1"], fill=fill_red),
        )


def _fill_para_columna(col_idx: int, config: ConfiguracionAnual, fills: dict[str, PatternFill]) -> PatternFill:
    headers = config.encabezados_visibles
    try:
        fin_geo = headers.index("VOTOS_EMITIDOS") + 1
    except ValueError:
        fin_geo = 8
    try:
        fin_participacion = headers.index("ABSTENCION") + 1
    except ValueError:
        fin_participacion = fin_geo + 2
    if col_idx <= fin_geo:
        return fills["geo"]
    if col_idx <= fin_participacion:
        return fills["participacion"]
    if col_idx < config.indice_inicio_partidos + 1:
        return fills["top"]
    if col_idx <= config.indice_nulos + 2:
        return fills["partidos"]
    return fills["validacion"]


def _aplicar_estilos_dataframe(ws, total_rows: int, total_cols: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}{max(total_rows + 1, 1)}"
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "base": PatternFill("solid", fgColor="D9EAF7"),
        "participacion": PatternFill("solid", fgColor="E2F0D9"),
        "top": PatternFill("solid", fgColor="FFF2CC"),
        "votos": PatternFill("solid", fgColor="FCE4D6"),
    }
    percent_headers = {"PARTICIPACION", "PARTICIPACION_PCN", "PARTICIPACION (%)"}
    text_headers = {"MUNICIPIO", "1ER_LUGAR", "2DO_LUGAR", "3ER_LUGAR", "1ER LUGAR", "2DO LUGAR", "3ER LUGAR"}

    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(1, col_idx)
        header = str(cell.value or "")
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.fill = _fill_dataframe(header, col_idx, fills)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(header) + 2, 10), 24)

    for row in ws.iter_rows(min_row=2, max_row=total_rows + 1):
        for cell in row:
            header = str(ws.cell(1, cell.column).value or "").upper()
            cell.border = border
            if header in percent_headers:
                cell.number_format = "0.00"
            elif header not in text_headers:
                cell.number_format = "#,##0"


def _fill_dataframe(header: str, col_idx: int, fills: dict[str, PatternFill]) -> PatternFill:
    header_norm = header.upper()
    if col_idx <= 4:
        return fills["base"]
    if "PARTICIPACION" in header_norm:
        return fills["participacion"]
    if "LUGAR" in header_norm or header_norm in {"VOTOS", "VOTOS.1", "VOTOS.2", "1ERO_VOTOS", "2DO_VOTOS", "3RO_VOTOS"}:
        return fills["top"]
    return fills["votos"]
